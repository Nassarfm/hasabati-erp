"""
scripts/generate_coa_template.py
══════════════════════════════════════════════════════════
يولّد ملف Excel جاهز يمكن للمستخدم تعبئته ورفعه.
يحتوي على:
  - ورقة "دليل الحسابات"  ← للمستخدم يعبّئها
  - ورقة "تعليمات"        ← شرح كل عمود
  - ورقة "مثال"           ← صف مثال لكل نوع حساب

تشغيل:
  python scripts/generate_coa_template.py
  → ينتج: coa_import_template.xlsx
══════════════════════════════════════════════════════════
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))


def generate_template(output_path: str = "coa_import_template.xlsx") -> str:
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("❌ openpyxl غير مثبت — شغّل: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════
    # Sheet 1: دليل الحسابات (main input sheet)
    # ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "دليل الحسابات"
    ws.sheet_view.rightToLeft = True

    # Colors
    HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark blue
    REQUIRED_FILL = PatternFill("solid", fgColor="FCE4D6")   # light orange
    OPTIONAL_FILL = PatternFill("solid", fgColor="E2EFDA")   # light green
    EXAMPLE_FILL  = PatternFill("solid", fgColor="FFF2CC")   # light yellow

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    BODY_FONT   = Font(size=10, name="Arial")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    COLUMNS = [
        ("code",            "كود الحساب *",         12, True),
        ("name_ar",         "اسم الحساب بالعربي *", 30, True),
        ("name_en",         "اسم الحساب بالإنجليزي",20, False),
        ("account_type",    "نوع الحساب *",          18, True),
        ("account_nature",  "طبيعة الحساب *",        16, True),
        ("level",           "المستوى *",             10, True),
        ("parent_code",     "كود الحساب الأب",       14, False),
        ("postable",        "قابل للترحيل *",        14, True),
        ("opening_balance", "الرصيد الافتتاحي",      16, False),
        ("notes",           "ملاحظات",               25, False),
    ]

    # Header row
    for col_idx, (_, label, width, required) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    # Data validation dropdowns
    type_validation = DataValidation(
        type="list",
        formula1='"asset,liability,equity,revenue,expense"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="قيمة غير صحيحة",
        error="يجب أن يكون: asset أو liability أو equity أو revenue أو expense",
    )
    nature_validation = DataValidation(
        type="list",
        formula1='"debit,credit"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="قيمة غير صحيحة",
        error="يجب أن يكون: debit أو credit",
    )
    postable_validation = DataValidation(
        type="list",
        formula1='"نعم,لا"',
        showDropDown=False,
    )
    level_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="4",
        showErrorMessage=True,
        errorTitle="مستوى غير صحيح",
        error="المستوى يجب أن يكون بين 1 و 4",
    )
    ws.add_data_validation(type_validation)
    ws.add_data_validation(nature_validation)
    ws.add_data_validation(postable_validation)
    ws.add_data_validation(level_validation)

    # Apply validations to data rows (2 to 1000)
    type_validation.sqref    = "D2:D1000"
    nature_validation.sqref  = "E2:E1000"
    postable_validation.sqref = "H2:H1000"
    level_validation.sqref   = "F2:F1000"

    # Example rows
    examples = [
        ("1100", "الأصول المتداولة", "Current Assets", "asset", "debit", 2, "", "لا", 0, "مجموعة"),
        ("1001", "الصندوق الرئيسي", "Main Cash", "asset", "debit", 4, "1100", "نعم", 5000, ""),
        ("2101", "ذمم الموردين", "Suppliers AP", "liability", "credit", 4, "2100", "نعم", 0, ""),
        ("4001", "مبيعات بضاعة", "Merchandise Sales", "revenue", "credit", 4, "4000", "نعم", 0, ""),
        ("6001", "رواتب الموظفين", "Employee Salaries", "expense", "debit", 4, "6200", "نعم", 0, ""),
    ]

    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = EXAMPLE_FILL
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right" if col_idx <= 2 else "center")

    # Empty template rows (light formatting)
    for row_idx in range(len(examples) + 2, 102):
        for col_idx, (_, _, _, required) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
            cell.border = BORDER
            cell.font = BODY_FONT

    # ══════════════════════════════════════════════════════
    # Sheet 2: تعليمات
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("تعليمات")
    ws2.sheet_view.rightToLeft = True

    instructions = [
        ("📋 تعليمات استيراد دليل الحسابات", True),
        ("", False),
        ("الأعمدة الإلزامية (باللون البرتقالي):", True),
        ("  • كود الحساب: رقم فريد مثل 1001 أو 4001", False),
        ("  • اسم الحساب بالعربي: الاسم الكامل", False),
        ("  • نوع الحساب: asset / liability / equity / revenue / expense", False),
        ("  • طبيعة الحساب: debit (مدين) أو credit (دائن)", False),
        ("  • المستوى: من 1 إلى 4", False),
        ("  • قابل للترحيل: نعم أو لا", False),
        ("", False),
        ("قواعد المستويات:", True),
        ("  المستوى 1: قسم رئيسي (أصول، التزامات...) — لا ترحيل", False),
        ("  المستوى 2: مجموعة رئيسية (أصول متداولة...) — لا ترحيل", False),
        ("  المستوى 3: حساب رئيسي (النقدية، المبيعات...) — لا ترحيل", False),
        ("  المستوى 4: حساب تحليلي — يقبل القيود المحاسبية ✓", False),
        ("", False),
        ("طبيعة الحساب الموصى بها:", True),
        ("  asset    → debit  (أصول)", False),
        ("  liability → credit (التزامات)", False),
        ("  equity   → credit (حقوق الملكية)", False),
        ("  revenue  → credit (إيرادات)", False),
        ("  expense  → debit  (مصروفات)", False),
        ("", False),
        ("⚠️  تنبيهات:", True),
        ("  • الأكواد يجب أن تكون فريدة لا تتكرر", False),
        ("  • كود الحساب الأب يجب أن يكون موجوداً في نفس الملف أو في قاعدة البيانات", False),
        ("  • الأرصدة الافتتاحية اختيارية — يمكن تركها صفراً", False),
        ("  • الصفوف الصفراء في ورقة 'دليل الحسابات' هي أمثلة فقط — احذفها قبل الرفع", False),
    ]

    for row_idx, (text, bold) in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=11, name="Arial")
        if bold and text:
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")

    ws2.column_dimensions["A"].width = 70

    # ══════════════════════════════════════════════════════
    # Sheet 3: دليل جاهز (مثال كامل)
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("دليل جاهز — مثال")
    ws3.sheet_view.rightToLeft = True
    ws3.cell(row=1, column=1, value="هذا الدليل موصى به — يمكنك نسخه إلى ورقة 'دليل الحسابات' مباشرة")
    ws3.cell(row=1, column=1).font = Font(bold=True, color="FF0000", size=12)

    # Save
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = generate_template()
    print(f"✅ تم إنشاء النموذج: {path}")
    print(f"   ارفعه على: POST /api/v1/accounting/coa/import")
