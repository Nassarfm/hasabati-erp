"""
app/modules/accounting/coa_import.py
══════════════════════════════════════════════════════════
استيراد دليل الحسابات من Excel أو CSV.
يدعم المستويات 1-5
══════════════════════════════════════════════════════════
"""
from __future__ import annotations

import csv
import io
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import structlog
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ValidationError
from app.core.tenant import CurrentUser
from app.modules.accounting.models import (
    AccountNature, AccountType, ChartOfAccount,
)
from app.modules.accounting.repository import COARepository

logger = structlog.get_logger(__name__)

VALID_TYPES    = {"asset", "liability", "equity", "revenue", "expense"}
VALID_NATURES  = {"debit", "credit"}
TRUTHY_VALUES  = {"نعم", "yes", "true", "1", "y", "صح"}

TYPE_MAP = {
    "أصول": "asset", "أصل": "asset",
    "التزامات": "liability", "التزام": "liability",
    "حقوق الملكية": "equity", "ملكية": "equity",
    "إيرادات": "revenue", "إيراد": "revenue",
    "مصروفات": "expense", "مصروف": "expense",
}
NATURE_MAP = {
    "مدين": "debit", "دائن": "credit",
}


@dataclass
class ImportError:
    row: int
    code: str
    field: str
    message: str


@dataclass
class ImportResult:
    total_rows: int = 0
    inserted: List[str] = field(default_factory=list)
    skipped: List[str] = field(default_factory=list)
    errors: List[ImportError] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def success_count(self) -> int:
        return len(self.inserted)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def has_errors(self) -> bool:
        return len(self.errors) > 0

    def to_dict(self) -> dict:
        return {
            "total_rows":    self.total_rows,
            "inserted":      self.success_count,
            "skipped":       len(self.skipped),
            "errors":        self.error_count,
            "inserted_codes": self.inserted[:50],
            "skipped_codes":  self.skipped[:50],
            "error_details": [
                {"row": e.row, "code": e.code, "field": e.field, "message": e.message}
                for e in self.errors[:100]
            ],
            "warnings": self.warnings[:20],
        }


def _normalize(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_bool(val: str) -> bool:
    return val.strip().lower() in TRUTHY_VALUES


def _parse_type(val: str) -> Optional[str]:
    v = val.strip().lower()
    if v in VALID_TYPES:
        return v
    return TYPE_MAP.get(val.strip())


def _parse_nature(val: str) -> Optional[str]:
    v = val.strip().lower()
    if v in VALID_NATURES:
        return v
    return NATURE_MAP.get(val.strip())


def _read_excel(content: bytes) -> Tuple[List[Dict], List[str]]:
    try:
        import openpyxl
    except ImportError:
        raise ValidationError("openpyxl غير مثبت — pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    sheet_name = wb.sheetnames[0]
    for name in wb.sheetnames:
        if "دليل" in name or "chart" in name.lower() or "coa" in name.lower():
            sheet_name = name
            break
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    header = [_normalize(h) for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        data_rows.append(dict(zip(header, row)))

    return data_rows, header


def _read_csv(content: bytes) -> Tuple[List[Dict], List[str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    headers = reader.fieldnames or []
    return rows, list(headers)


COLUMN_ALIASES = {
    "code":            ["code", "كود", "كود الحساب", "الكود", "رقم الحساب"],
    "name_ar":         ["name_ar", "name", "اسم", "اسم الحساب", "الاسم", "اسم الحساب بالعربي"],
    "name_en":         ["name_en", "english", "الاسم الإنجليزي", "اسم الحساب بالإنجليزي"],
    "account_type":    ["account_type", "type", "نوع", "نوع الحساب"],
    "account_nature":  ["account_nature", "nature", "طبيعة", "طبيعة الحساب"],
    "level":           ["level", "مستوى", "المستوى"],
    "parent_code":     ["parent_code", "parent", "أب", "الحساب الأب", "كود الحساب الأب"],
    "postable":        ["postable", "قابل للترحيل", "ترحيل"],
    "opening_balance": ["opening_balance", "رصيد افتتاحي", "الرصيد الافتتاحي", "رصيد"],
    "notes":           ["notes", "ملاحظات"],
    "function_type":   ["function_type", "function", "bs/pl", "نوع القائمة", "func"],
    "grp":             ["grp", "group", "المجموعة", "مجموعة"],
    "sub_group":       ["sub_group", "sub-group", "subgroup", "المجموعة الفرعية", "sub group"],
    "cash_flow_type":  ["cash_flow_type", "cash flow type", "cash_flow", "التدفق النقدي", "cash flow"],
    "dimension_required": ["dimension_required", "dimension required", "يتطلب بعد", "dimension"],
}


def _map_columns(header: List[str]) -> Dict[str, str]:
    mapping = {}
    for field_name, aliases in COLUMN_ALIASES.items():
        for h in header:
            if h.strip().lower() in [a.lower() for a in aliases]:
                mapping[h.strip()] = field_name
                break
    return mapping


class COAImportService:
    def __init__(self, db: AsyncSession, user: CurrentUser) -> None:
        self.db = db
        self.user = user
        self._repo = COARepository(db, user.tenant_id)

    async def import_file(
        self,
        content: bytes,
        filename: str,
        dry_run: bool = False,
    ) -> ImportResult:
        self.user.require("can_manage_coa")

        if filename.lower().endswith(".csv"):
            raw_rows, header = _read_csv(content)
        elif filename.lower().endswith((".xlsx", ".xls")):
            raw_rows, header = _read_excel(content)
        else:
            raise ValidationError("صيغة الملف غير مدعومة — يُقبل: .xlsx أو .csv فقط")

        if not raw_rows:
            raise ValidationError("الملف فارغ أو لا يحتوي على بيانات")

        col_map = _map_columns(header)
        if "code" not in col_map.values():
            raise ValidationError("لم يتم العثور على عمود الكود — يجب أن يكون العنوان: 'code' أو 'كود الحساب'")
        if "name_ar" not in col_map.values():
            raise ValidationError("لم يتم العثور على عمود الاسم العربي")

        result = ImportResult(total_rows=len(raw_rows))

        existing_accounts = await self._repo.list_active()
        existing_codes = {a.code for a in existing_accounts}

        file_codes: Dict[str, int] = {}

        parsed_rows = []
        for row_idx, raw in enumerate(raw_rows, 2):
            def get(field: str) -> str:
                for raw_col, mapped_field in col_map.items():
                    if mapped_field == field:
                        return _normalize(raw.get(raw_col, ""))
                return ""

            code     = get("code")
            name_ar  = get("name_ar")
            name_en  = get("name_en")
            acc_type = get("account_type")
            nature   = get("account_nature")
            level_s  = get("level")
            parent   = get("parent_code")
            postable = get("postable")
            ob_str   = get("opening_balance")
            notes              = get("notes")
            function_type      = get("function_type") or "BS"
            grp                = get("grp") or ""
            sub_group          = get("sub_group") or ""
            cash_flow_type_raw = get("cash_flow_type") or "none"
            dim_raw            = get("dimension_required")

            # تحويل cash_flow_type
            cf_map = {
                "operating activities": "operating", "operating": "operating", "أنشطة تشغيلية": "operating",
                "investing activities": "investing", "investing": "investing", "أنشطة استثمارية": "investing",
                "financing activities": "financing", "financing": "financing", "أنشطة تمويلية": "financing",
                "none": "none", "لا ينطبق": "none", "": "none",
            }
            cash_flow_type = cf_map.get(cash_flow_type_raw.strip().lower(), "none")

            # تحويل dimension_required
            dimension_required = dim_raw.strip().lower() in {"نعم", "yes", "true", "1", "y"}

            # تحويل function_type
            ft_map = {"bs": "BS", "pl": "PL", "bs/pl": "BS/PL", "": "BS"}
            function_type = ft_map.get(function_type.strip().lower(), function_type.strip().upper() or "BS")

            row_errors = []

            if not code:
                row_errors.append(ImportError(row_idx, "", "code", "كود الحساب مطلوب"))
                result.errors.extend(row_errors)
                continue

            if not name_ar:
                row_errors.append(ImportError(row_idx, code, "name_ar", "اسم الحساب بالعربي مطلوب"))

            parsed_type = _parse_type(acc_type)
            if not parsed_type:
                row_errors.append(ImportError(
                    row_idx, code, "account_type",
                    f"نوع الحساب '{acc_type}' غير صحيح — يجب: asset/liability/equity/revenue/expense"
                ))

            parsed_nature = _parse_nature(nature)
            if not parsed_nature:
                row_errors.append(ImportError(
                    row_idx, code, "account_nature",
                    f"طبيعة الحساب '{nature}' غير صحيحة — يجب: debit أو credit"
                ))

            # ✅ رفع الحد إلى 5 مستويات
            try:
                level = int(float(level_s)) if level_s else 1
                if level not in (1, 2, 3, 4, 5):
                    raise ValueError()
            except (ValueError, TypeError):
                level = 1
                row_errors.append(ImportError(
                    row_idx, code, "level", f"المستوى '{level_s}' غير صحيح — يجب 1 إلى 5"
                ))

            if code in file_codes:
                row_errors.append(ImportError(
                    row_idx, code, "code",
                    f"الكود '{code}' مكرر في الملف (ظهر أول مرة في السطر {file_codes[code]})"
                ))
            else:
                file_codes[code] = row_idx

            try:
                opening_balance = float(ob_str) if ob_str else 0.0
            except ValueError:
                opening_balance = 0.0
                result.warnings.append(f"السطر {row_idx}: رصيد افتتاحي غير صحيح '{ob_str}' — تم تعيينه 0")

            if code in existing_codes:
                result.skipped.append(code)
                continue

            if row_errors:
                result.errors.extend(row_errors)
                continue

            parsed_rows.append({
                "row_idx":            row_idx,
                "code":               code,
                "name_ar":            name_ar,
                "name_en":            name_en or None,
                "account_type":       parsed_type,
                "account_nature":     parsed_nature,
                "level":              level,
                "parent_code":        parent or None,
                "postable":           _parse_bool(postable),
                "opening_balance":    opening_balance,
                "notes":              notes or None,
                "function_type":      function_type,
                "grp":                grp or None,
                "sub_group":          sub_group or None,
                "cash_flow_type":     cash_flow_type,
                "dimension_required": dimension_required,
            })

        if result.has_errors:
            logger.warning("coa_import_validation_failed", errors=result.error_count, tenant=str(self.user.tenant_id))
            return result

        if dry_run:
            result.inserted = [r["code"] for r in parsed_rows]
            return result

        now = datetime.now(timezone.utc)

        code_to_uuid: Dict[str, uuid.UUID] = {a.code: a.id for a in existing_accounts}
        for r in parsed_rows:
            if r["code"] not in code_to_uuid:
                code_to_uuid[r["code"]] = uuid.uuid4()

        for r in parsed_rows:
            parent_id = None
            if r["parent_code"]:
                parent_id = code_to_uuid.get(r["parent_code"])
                if not parent_id:
                    result.warnings.append(
                        f"السطر {r['row_idx']}: كود الحساب الأب '{r['parent_code']}' غير موجود — تم تعيينه بدون أب"
                    )

            acc = ChartOfAccount(
                id=code_to_uuid[r["code"]],
                tenant_id=self.user.tenant_id,
                code=r["code"],
                name_ar=r["name_ar"],
                name_en=r["name_en"],
                account_type=r["account_type"],
                account_nature=r["account_nature"],
                level=r["level"],
                parent_id=parent_id,
                postable=r["postable"],
                allow_direct_posting=r["postable"],
                is_active=True,
                opening_balance=r["opening_balance"],
                function_type=r.get("function_type"),
                grp=r.get("grp"),
                sub_group=r.get("sub_group"),
                cash_flow_type=r.get("cash_flow_type", "none"),
                dimension_required=r.get("dimension_required", False),
                created_by=self.user.email,
                created_at=now,
                updated_at=now,
            )
            self.db.add(acc)
            result.inserted.append(r["code"])

        await self.db.flush()

        logger.info("coa_import_success", inserted=result.success_count, skipped=len(result.skipped), tenant=str(self.user.tenant_id))
        return result
