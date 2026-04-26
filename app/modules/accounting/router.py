"""
app/modules/accounting/router.py
══════════════════════════════════════════════════════════
Accounting Module API

Chart of Accounts:
  GET    /accounting/coa
  POST   /accounting/coa
  GET    /accounting/coa/{id}
  PUT    /accounting/coa/{id}
  DELETE /accounting/coa/reset
  POST   /accounting/coa/import
  GET    /accounting/coa/template
  POST   /accounting/coa/seed

Journal Entries:
  POST   /accounting/je
  GET    /accounting/je
  GET    /accounting/je/{id}
  PUT    /accounting/je/{id}
  POST   /accounting/je/{id}/submit
  POST   /accounting/je/{id}/approve
  POST   /accounting/je/{id}/reject
  POST   /accounting/je/{id}/post
  POST   /accounting/je/{id}/reverse

Recurring Entries:
  POST   /accounting/recurring/preview
  POST   /accounting/recurring
  GET    /accounting/recurring
  GET    /accounting/recurring/{id}
  POST   /accounting/recurring/{id}/post-pending
  POST   /accounting/recurring/instances/{id}/skip
  PATCH  /accounting/recurring/{id}/status
  DELETE /accounting/recurring/{id}

Fiscal Periods:
  GET    /accounting/fiscal-locks
  POST   /accounting/fiscal-locks
  DELETE /accounting/fiscal-locks/{id}

Reports:
  GET    /accounting/ledger/{code}
  GET    /accounting/trial-balance
  POST   /accounting/rebuild-balances
  GET    /accounting/dashboard
══════════════════════════════════════════════════════════
"""
from __future__ import annotations

import uuid
from datetime import date
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.response import created, ok, paginated
from app.core.tenant import CurrentUser, get_current_user
from app.db.session import get_db
from app.modules.accounting.schemas import (
    COAAccountCreate, COAAccountUpdate, JournalEntryCreate,
    LockPeriodRequest, PostJERequest, ReverseJERequest,
)
from app.modules.accounting.service import AccountingService

router = APIRouter(prefix="/accounting", tags=["المحاسبة"])


def _svc(
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(get_current_user),
) -> AccountingService:
    return AccountingService(db, user)


# ══════════════════════════════════════════════════════════
# Chart of Accounts
# ══════════════════════════════════════════════════════════
@router.get("/coa", summary="قائمة الحسابات")
async def list_coa(svc: AccountingService = Depends(_svc)):
    accounts = await svc.list_accounts()
    return ok(data=[a.to_dict() for a in accounts], message=f"{len(accounts)} حساب")


@router.post("/coa", status_code=201, summary="إنشاء حساب جديد")
async def create_account(data: COAAccountCreate, svc: AccountingService = Depends(_svc)):
    acc = await svc.create_account(data)
    return created(data=acc.to_dict(), message=f"تم إنشاء الحساب {data.code}")


@router.get("/coa/{account_id}", summary="تفاصيل حساب")
async def get_account(account_id: uuid.UUID, svc: AccountingService = Depends(_svc)):
    acc = await svc._coa_repo.get_or_raise(account_id)
    return ok(data=acc.to_dict())


@router.put("/coa/{account_id}", summary="تعديل حساب")
async def update_account(
    account_id: uuid.UUID,
    data: COAAccountUpdate,
    svc: AccountingService = Depends(_svc),
):
    acc = await svc.update_account(account_id, data)
    return ok(data=acc.to_dict(), message=f"تم تعديل الحساب {acc.code}")


@router.delete("/coa/reset", summary="إعادة تهيئة دليل الحسابات")
async def reset_coa(svc: AccountingService = Depends(_svc)):
    result = await svc.reset_coa()
    return ok(data=result, message=result["message"])


@router.post("/coa/import", status_code=201, summary="استيراد دليل الحسابات", response_model=None)
async def import_coa(
    file:    UploadFile = File(...),
    dry_run: bool       = Query(default=False),
    db:      AsyncSession  = Depends(get_db),
    user:    CurrentUser   = Depends(get_current_user),
):
    from app.modules.accounting.coa_import import COAImportService
    content = await file.read()
    if not content:
        from app.core.exceptions import ValidationError
        raise ValidationError("الملف فارغ")
    svc = COAImportService(db, user)
    result = await svc.import_file(content=content, filename=file.filename or "upload.xlsx", dry_run=dry_run)
    if result.has_errors:
        return ok(data=result.to_dict(), message=f"❌ يوجد {result.error_count} خطأ")
    action = "معاينة" if dry_run else "استيراد"
    return created(data=result.to_dict(), message=f"✅ {action} ناجح — {result.success_count} حساب")


@router.get("/coa/template", summary="تحميل نموذج Excel", response_class=StreamingResponse)
async def download_coa_template():
    import io
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
    except ImportError:
        from app.core.exceptions import ValidationError
        raise ValidationError("openpyxl غير مثبت على الخادم")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "دليل الحسابات"
    ws.sheet_view.rightToLeft = True
    H_FILL = PatternFill("solid", fgColor="1F3864")
    EX_FILL = PatternFill("solid", fgColor="FFF2CC")
    R_FILL = PatternFill("solid", fgColor="FCE4D6")
    O_FILL = PatternFill("solid", fgColor="E2EFDA")
    H_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    B_FONT = Font(size=10, name="Arial")
    THIN = Side(style="thin", color="BFBFBF")
    BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    COLS = [
        ("code","كود الحساب *",12,True),("name_ar","اسم الحساب بالعربي *",30,True),
        ("name_en","اسم الحساب بالإنجليزي",20,False),("account_type","نوع الحساب *",18,True),
        ("account_nature","طبيعة الحساب *",16,True),("level","المستوى (1-4) *",12,True),
        ("parent_code","كود الحساب الأب",14,False),("postable","قابل للترحيل (نعم/لا) *",18,True),
        ("opening_balance","الرصيد الافتتاحي",16,False),("notes","ملاحظات",20,False),
    ]
    for i, (_, label, width, req) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = H_FONT; c.fill = H_FILL; c.alignment = CTR; c.border = BRD
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    dv_type   = DataValidation(type="list", formula1='"asset,liability,equity,revenue,expense"', showDropDown=False)
    dv_nature = DataValidation(type="list", formula1='"debit,credit"', showDropDown=False)
    dv_post   = DataValidation(type="list", formula1='"نعم,لا"', showDropDown=False)
    for dv in [dv_type, dv_nature, dv_post]: ws.add_data_validation(dv)
    dv_type.sqref = "D2:D500"; dv_nature.sqref = "E2:E500"; dv_post.sqref = "H2:H500"
    EXAMPLES = [
        ("11","الأصول المتداولة","Current Assets","asset","debit",2,"1","لا",0,""),
        ("1001","الصندوق الرئيسي","Main Cash","asset","debit",4,"11","نعم",5000,""),
        ("2101","ذمم الموردين","Suppliers AP","liability","credit",4,"21","نعم",0,""),
        ("4001","مبيعات بضاعة","Merchandise Sales","revenue","credit",4,"41","نعم",0,""),
        ("5001","تكلفة البضاعة المباعة","COGS","expense","debit",4,"51","نعم",0,""),
        ("6001","رواتب الموظفين","Employee Salaries","expense","debit",4,"6200","نعم",0,""),
    ]
    for ri, row in enumerate(EXAMPLES, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = EX_FILL; c.font = B_FONT; c.border = BRD
            c.alignment = Alignment(horizontal="right" if ci <= 2 else "center")
    for ri in range(len(EXAMPLES) + 2, 102):
        for ci, (_, _, _, req) in enumerate(COLS, 1):
            c = ws.cell(row=ri, column=ci)
            c.fill = R_FILL if req else O_FILL; c.border = BRD; c.font = B_FONT
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="coa_template.xlsx"'})


@router.post("/coa/seed", status_code=201, summary="تحميل دليل الحسابات الجاهز", response_model=None)
async def seed_default_coa(db: AsyncSession = Depends(get_db), user: CurrentUser = Depends(get_current_user)):
    from app.modules.accounting.coa_import import COAImportService
    from scripts.seed_coa import COA
    user.require("can_manage_coa")
    import csv, io
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["code","name_ar","name_en","account_type","account_nature","level","parent_code","postable","opening_balance"])
    writer.writeheader()
    for acc in COA:
        writer.writerow({"code":acc.code,"name_ar":acc.name_ar,"name_en":acc.name_en,"account_type":acc.account_type,"account_nature":acc.account_nature,"level":acc.level,"parent_code":acc.parent_code or "","postable":"نعم" if acc.postable else "لا","opening_balance":acc.opening_balance})
    svc = COAImportService(db, user)
    result = await svc.import_file(content=buf.getvalue().encode("utf-8"), filename="seed.csv", dry_run=False)
    return created(data=result.to_dict(), message=f"✅ تم تحميل دليل الحسابات — {result.success_count} حساب جديد")


# ══════════════════════════════════════════════════════════
# Journal Entries
# ══════════════════════════════════════════════════════════
@router.post("/je", status_code=201, summary="إنشاء قيد محاسبي (draft)")
async def create_je(
    data: JournalEntryCreate,
    db:   AsyncSession = Depends(get_db),
    user: CurrentUser  = Depends(get_current_user),
    svc: AccountingService = Depends(_svc),
):
    from sqlalchemy import text as _txt

    je = await svc.create_draft_je(data)

    # ── حفظ party_id لكل سطر يحتوي متعاملاً ──────────────
    lines_with_party = [l for l in data.lines if getattr(l, "party_id", None)]
    if lines_with_party:
        try:
            # جلب je_lines المُنشأة بالترتيب
            res = await db.execute(_txt("""
                SELECT id, line_order FROM je_lines
                WHERE journal_entry_id = :je_id
                ORDER BY line_order
            """), {"je_id": str(je.id)})
            saved_lines = res.fetchall()
            for idx, saved in enumerate(saved_lines):
                if idx < len(data.lines):
                    src = data.lines[idx]
                    _pid  = getattr(src, "party_id",   None)
                    _role = getattr(src, "party_role",  None)
                    if _pid:
                        _pid_s = str(_pid)
                        _role_v = _role or "other"
                        _name_v = getattr(src, "party_name", None)
                        try:
                            # محاولة مع party_name
                            await db.execute(_txt(f"""
                                UPDATE je_lines
                                SET party_id   = '{_pid_s}'::uuid,
                                    party_role = :prole,
                                    party_name = :pname
                                WHERE id = :line_id
                            """), {"prole": _role_v, "pname": _name_v, "line_id": str(saved[0])})
                        except Exception:
                            try:
                                # محاولة بدون party_name إذا كان العمود غير موجود
                                await db.execute(_txt(f"""
                                    UPDATE je_lines
                                    SET party_id   = '{_pid_s}'::uuid,
                                        party_role = :prole
                                    WHERE id = :line_id
                                """), {"prole": _role_v, "line_id": str(saved[0])})
                            except Exception as e2:
                                import logging
                                logging.getLogger(__name__).warning(f"party_save_failed: {e2}")
            await db.commit()
        except Exception as e:
            # لا نوقف الإنشاء إذا فشل حفظ الـ party
            import logging; logging.getLogger(__name__).warning(f"party_save_skipped: {e}")

    return created(
        data={"id":str(je.id),"serial":je.serial,"status":je.status,
              "total_debit":float(je.total_debit),"total_credit":float(je.total_credit)},
        message=f"تم إنشاء القيد {je.serial}"
    )


@router.get("/je", summary="قائمة القيود المحاسبية")
async def list_je(
    status:      Optional[str]     = Query(None),
    je_type:     Optional[str]     = Query(None),
    date_from:   Optional[date]    = Query(None),
    date_to:     Optional[date]    = Query(None),
    fiscal_year: Optional[int]     = Query(None),
    search:      Optional[str]     = Query(None),
    created_by:  Optional[str]     = Query(None),
    min_amount:  Optional[Decimal] = Query(None),
    max_amount:  Optional[Decimal] = Query(None),
    limit:       int               = Query(20, ge=1, le=1000),
    offset:      int               = Query(0, ge=0),
    db:  AsyncSession  = Depends(get_db),
    user: CurrentUser  = Depends(get_current_user),
    svc: AccountingService = Depends(_svc),
):
    from sqlalchemy import text as _txt

    items, total = await svc.list_je(
        status=status,
        je_type=je_type,
        date_from=date_from,
        date_to=date_to,
        fiscal_year=fiscal_year,
        offset=offset,
        limit=limit,
    )
    # فلترة client-side
    if search:
        s = search.lower()
        items = [j for j in items if s in (j.serial or '').lower() or s in (j.description or '').lower()]
    if created_by:
        items = [j for j in items if created_by.lower() in (j.created_by or '').lower()]
    if min_amount is not None:
        items = [j for j in items if float(j.total_debit or 0) >= float(min_amount) or float(j.total_credit or 0) >= float(min_amount)]
    if max_amount is not None:
        items = [j for j in items if float(j.total_debit or 0) <= float(max_amount) or float(j.total_credit or 0) <= float(max_amount)]

    # ── جلب بيانات المتعامل الأول لكل قيد ─────────────────
    je_ids = [str(j.id) for j in items]
    party_map: dict = {}
    if je_ids:
        try:
            # نبني IN clause يدوياً لتجنب مشكلة asyncpg مع ANY
            placeholders = ','.join([f"'{_id}'" for _id in je_ids])
            res = await db.execute(_txt(f"""
                SELECT DISTINCT ON (jl.journal_entry_id)
                    jl.journal_entry_id AS je_id,
                    jl.party_id,
                    jl.party_role,
                    p.party_name_ar AS party_name,
                    p.party_code
                FROM je_lines jl
                LEFT JOIN parties p ON p.id::text = jl.party_id::text
                WHERE jl.journal_entry_id::text IN ({placeholders})
                  AND jl.party_id IS NOT NULL
                ORDER BY jl.journal_entry_id, jl.line_order
            """))
            for row in res.mappings().fetchall():
                party_map[str(row["je_id"])] = {
                    "party_id":   str(row["party_id"]) if row["party_id"] else None,
                    "party_name": row["party_name"] or row["party_code"] or None,
                    "party_role": row["party_role"] or None,
                }
        except Exception as _e:
            import logging; logging.getLogger(__name__).warning(f"party_map_failed: {_e}")

    result = []
    for j in items:
        d = j.to_dict()
        pinfo = party_map.get(str(j.id), {})
        d["first_party_id"]   = pinfo.get("party_id")
        d["first_party_name"] = pinfo.get("party_name")
        d["first_party_role"] = pinfo.get("party_role")
        result.append(d)

    return {
        "data":        result,
        "total_count": total,
        "total":       total,
        "count":       total,
        "limit":       limit,
        "offset":      offset,
    }


@router.get("/je/{je_id}", summary="تفاصيل قيد محاسبي")
async def get_je(
    je_id: uuid.UUID,
    db:  AsyncSession  = Depends(get_db),
    user: CurrentUser  = Depends(get_current_user),
    svc: AccountingService = Depends(_svc),
):
    from sqlalchemy import text as _txt

    je   = await svc.get_je(je_id)
    data = je.to_dict()

    # ── جلب party_id / party_name لكل سطر ─────────────────
    try:
        res = await db.execute(_txt("""
            SELECT
                jl.id,
                jl.party_id,
                jl.party_role,
                COALESCE(p.party_name_ar, jl.party_name) AS party_name,
                p.party_code
            FROM je_lines jl
            LEFT JOIN parties p ON p.id::text = jl.party_id::text
            WHERE jl.journal_entry_id = :je_id
            ORDER BY jl.line_order
        """), {"je_id": str(je_id)})
        party_line_map = {}
        for row in res.mappings().fetchall():
            party_line_map[str(row["id"])] = {
                "party_id":   str(row["party_id"]) if row["party_id"] else None,
                "party_name": row["party_name"] or row["party_code"] or None,
                "party_role": row["party_role"] or None,
            }
    except Exception:
        party_line_map = {}

    lines = []
    for l in je.lines:
        ld = l.to_dict()
        pinfo = party_line_map.get(str(l.id), {})
        ld["party_id"]   = pinfo.get("party_id")   or getattr(l, "party_id",   None)
        ld["party_name"] = pinfo.get("party_name") or None
        ld["party_role"] = pinfo.get("party_role") or getattr(l, "party_role", None)
        lines.append(ld)

    data["lines"] = lines
    return ok(data=data)


@router.put("/je/{je_id}", summary="تعديل قيد مسودة")
async def update_je(
    je_id: uuid.UUID,
    data:  JournalEntryCreate,
    db:    AsyncSession = Depends(get_db),
    user:  CurrentUser  = Depends(get_current_user),
    svc:   AccountingService = Depends(_svc),
):
    from sqlalchemy import text as _txt

    je = await svc.update_draft_je(je_id, data)

    # ── تحديث party_id بعد تعديل المسودة ─────────────────
    lines_with_party = [l for l in data.lines if getattr(l, "party_id", None)]
    if lines_with_party:
        try:
            res = await db.execute(_txt("""
                SELECT id, line_order FROM je_lines
                WHERE journal_entry_id = :je_id ORDER BY line_order
            """), {"je_id": str(je.id)})
            saved_lines = res.fetchall()
            for idx, saved in enumerate(saved_lines):
                if idx < len(data.lines):
                    src = data.lines[idx]
                    _pid  = getattr(src, "party_id",  None)
                    _role = getattr(src, "party_role", None)
                    if _pid:
                        _pid_s = str(_pid)
                        _role_v = _role or "other"
                        _name_v = getattr(src, "party_name", None)
                        try:
                            await db.execute(_txt(f"""
                                UPDATE je_lines
                                SET party_id='{_pid_s}'::uuid, party_role=:prole, party_name=:pname
                                WHERE id=:line_id
                            """), {"prole":_role_v, "pname":_name_v, "line_id":str(saved[0])})
                        except Exception:
                            try:
                                await db.execute(_txt(f"""
                                    UPDATE je_lines
                                    SET party_id='{_pid_s}'::uuid, party_role=:prole
                                    WHERE id=:line_id
                                """), {"prole":_role_v, "line_id":str(saved[0])})
                            except Exception as e2:
                                import logging
                                logging.getLogger(__name__).warning(f"party_update_failed: {e2}")
            await db.commit()
        except Exception as e:
            import logging; logging.getLogger(__name__).warning(f"party_update_skipped: {e}")

    return ok(data={"id":str(je.id),"serial":je.serial,"status":je.status},
              message=f"تم تعديل القيد {je.serial}")


@router.post("/je/{je_id}/submit", summary="إرسال للمراجعة")
async def submit_je(je_id: uuid.UUID, svc: AccountingService = Depends(_svc)):
    je = await svc.submit_je(je_id)
    return ok(data={"id":str(je.id),"status":je.status}, message="تم إرسال القيد للمراجعة")


@router.post("/je/{je_id}/approve", summary="الموافقة على القيد")
async def approve_je(je_id: uuid.UUID, svc: AccountingService = Depends(_svc)):
    je = await svc.approve_je(je_id)
    return ok(data={"id":str(je.id),"status":je.status}, message="تمت الموافقة وترحيل القيد")


@router.post("/je/{je_id}/reject", summary="رفض القيد")
async def reject_je(je_id: uuid.UUID, body: dict = Body(default={}), svc: AccountingService = Depends(_svc)):
    note = body.get("note", "")
    je = await svc.reject_je(je_id, note)
    return ok(data={"id":str(je.id),"status":je.status}, message="تم رفض القيد")


@router.post("/je/{je_id}/post", summary="ترحيل قيد محاسبي")
async def post_je(
    je_id:  uuid.UUID,
    body:   PostJERequest,
    db:     AsyncSession = Depends(get_db),
    user:   CurrentUser  = Depends(get_current_user),
    svc: AccountingService = Depends(_svc),
):
    from sqlalchemy import text as _txt

    # ── 1. احفظ party_id من je_lines قبل الترحيل (قد تُعاد كتابتها) ──
    party_map: dict = {}
    try:
        res = await db.execute(_txt("""
            SELECT line_order, party_id, party_role, party_name
            FROM je_lines
            WHERE journal_entry_id = :jid
              AND party_id IS NOT NULL
            ORDER BY line_order
        """), {"jid": str(je_id)})
        for row in res.mappings().fetchall():
            party_map[int(row["line_order"])] = {
                "party_id":   str(row["party_id"]) if row["party_id"] else None,
                "party_role": row["party_role"],
                "party_name": row["party_name"],
            }
    except Exception:
        pass  # العمود قد لا يكون موجوداً بعد

    # ── 2. نفّذ الترحيل ──────────────────────────────────
    je = await svc.post_je(je_id, force=body.force)

    # ── 3. أعد حفظ party_id على الـ je_lines الجديدة بعد الترحيل ──
    if party_map:
        try:
            res = await db.execute(_txt("""
                SELECT id, line_order FROM je_lines
                WHERE journal_entry_id = :jid
                ORDER BY line_order
            """), {"jid": str(je.id)})
            saved_lines = res.mappings().fetchall()
            for saved in saved_lines:
                order = int(saved["line_order"])
                pinfo = party_map.get(order)
                if not pinfo or not pinfo["party_id"]:
                    continue
                pid = pinfo["party_id"]
                try:
                    await db.execute(_txt(f"""
                        UPDATE je_lines
                        SET party_id   = '{pid}'::uuid,
                            party_role = :prole,
                            party_name = :pname
                        WHERE id = :lid
                    """), {
                        "prole": pinfo["party_role"] or "other",
                        "pname": pinfo["party_name"],
                        "lid":   str(saved["id"]),
                    })
                except Exception:
                    try:
                        await db.execute(_txt(f"""
                            UPDATE je_lines
                            SET party_id = '{pid}'::uuid, party_role = :prole
                            WHERE id = :lid
                        """), {"prole": pinfo["party_role"] or "other", "lid": str(saved["id"])})
                    except Exception as e2:
                        import logging
                        logging.getLogger(__name__).warning(f"post_je_party_restore_failed: {e2}")
            await db.commit()
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"post_je_party_map_failed: {e}")

    return ok(
        data={
            "id":           str(je.id),
            "serial":       je.serial,
            "status":       je.status,
            "posted_at":    str(je.posted_at),
            "total_debit":  float(je.total_debit),
            "total_credit": float(je.total_credit),
        },
        message="تم ترحيل القيد " + je.serial
    )


@router.post("/je/{je_id}/reverse", summary="عكس قيد محاسبي")
async def reverse_je(je_id: uuid.UUID, data: ReverseJERequest, svc: AccountingService = Depends(_svc)):
    result = await svc.reverse_je(je_id, data)
    return ok(data=result, message=f"✅ تم عكس القيد — {result.get('je_serial')}")


# ══════════════════════════════════════════════════════════
# Recurring Entries — القيود المتكررة
# ══════════════════════════════════════════════════════════
from app.modules.accounting.recurring_router import router as recurring_router
router.include_router(recurring_router)


# ══════════════════════════════════════════════════════════
# Fiscal Periods & Locks
# ══════════════════════════════════════════════════════════
@router.get("/fiscal-locks", summary="أقفال الفترات المالية")
async def list_locks(svc: AccountingService = Depends(_svc)):
    locks = await svc.list_locks()
    return ok(data=locks)


@router.post("/fiscal-locks", status_code=201, summary="قفل فترة مالية")
async def lock_period(data: LockPeriodRequest, svc: AccountingService = Depends(_svc)):
    result = await svc.lock_period(data)
    lock_label = "صارم" if data.lock_type == "hard" else "ناعم"
    return created(data=result, message=f"تم قفل الفترة {data.fiscal_year}" + (f"/{data.fiscal_month:02d}" if data.fiscal_month else "") + f" — قفل {lock_label}")


@router.delete("/fiscal-locks/{lock_id}", summary="فك قفل فترة مالية")
async def unlock_period(lock_id: uuid.UUID, svc: AccountingService = Depends(_svc)):
    result = await svc.unlock_period(lock_id)
    return ok(data=result, message="تم فك القفل")


# ══════════════════════════════════════════════════════════
# Reports
# ══════════════════════════════════════════════════════════
@router.get("/trial-balance", summary="ميزان المراجعة")
async def trial_balance(
    fiscal_year:  int           = Query(..., ge=2000, le=2100),
    fiscal_month: Optional[int] = Query(None, ge=1, le=12),
    svc: AccountingService = Depends(_svc),
):
    result = await svc.get_trial_balance(fiscal_year, fiscal_month)
    return ok(data=result)


@router.post("/rebuild-balances", summary="إعادة بناء الأرصدة")
async def rebuild_balances(
    fiscal_year: int = Query(..., description="السنة المالية"),
    svc: AccountingService = Depends(_svc),
):
    result = await svc.rebuild_balances(fiscal_year)
    return ok(data=result)


@router.get("/ledger/{account_code}", summary="كشف حساب — الأستاذ العام")
async def account_ledger(
    account_code: str,
    date_from:    Optional[date] = Query(None),
    date_to:      Optional[date] = Query(None),
    svc: AccountingService = Depends(_svc),
):
    from sqlalchemy import select
    from app.modules.accounting.models import JournalEntryLine, JournalEntry
    from decimal import Decimal

    q = (
        select(JournalEntryLine, JournalEntry)
        .join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id)
        .where(JournalEntryLine.tenant_id == svc.user.tenant_id)
        .where(JournalEntryLine.account_code == account_code)
        .where(JournalEntry.status == "posted")
    )
    if date_from: q = q.where(JournalEntry.entry_date >= date_from)
    if date_to:   q = q.where(JournalEntry.entry_date <= date_to)
    q = q.order_by(JournalEntry.entry_date, JournalEntry.serial)

    result = await svc.db.execute(q)
    rows = result.all()

    running = Decimal("0")
    lines = []
    for line, je in rows:
        running += line.debit - line.credit
        lines.append({
            "je_serial":        je.serial,
            "entry_date":       str(je.entry_date),
            "description":      line.description or je.description,
            "debit":            float(line.debit),
            "credit":           float(line.credit),
            "running_balance":  float(running),
            "je_type":          je.je_type,
            "created_by":       je.created_by if hasattr(je, "created_by") else "",
            "source_doc_number": je.source_doc_number,
        })

    # حساب رصيد الافتتاح (كل الحركات قبل date_from)
    opening = Decimal("0")
    if date_from:
        q2 = (
            select(JournalEntryLine, JournalEntry)
            .join(JournalEntry, JournalEntryLine.journal_entry_id == JournalEntry.id)
            .where(JournalEntryLine.tenant_id == svc.user.tenant_id)
            .where(JournalEntryLine.account_code == account_code)
            .where(JournalEntry.status == "posted")
            .where(JournalEntry.entry_date < date_from)
        )
        r2 = await svc.db.execute(q2)
        for line2, _ in r2.all():
            opening += line2.debit - line2.credit

    return ok(data={
        "account_code":    account_code,
        "opening_balance": float(opening),
        "closing_balance": float(opening + running),
        "transactions":    lines,
        "total_debit":     float(sum(l["debit"]  for l in lines)),
        "total_credit":    float(sum(l["credit"] for l in lines)),
    })


@router.get("/dashboard", summary="ملخص محاسبي")
async def accounting_dashboard(
    fiscal_year: int = Query(default=2024),
    svc: AccountingService = Depends(_svc),
):
    from sqlalchemy import func, select
    from app.modules.accounting.models import JournalEntry
    result = await svc.db.execute(
        select(JournalEntry.status, func.count())
        .where(JournalEntry.tenant_id == svc.user.tenant_id)
        .where(JournalEntry.fiscal_year == fiscal_year)
        .group_by(JournalEntry.status)
    )
    counts = {row[0]: row[1] for row in result.all()}
    return ok(data={
        "fiscal_year":   fiscal_year,
        "je_draft":      counts.get("draft", 0),
        "je_posted":     counts.get("posted", 0),
        "je_reversed":   counts.get("reversed", 0),
        "active_locks":  len(await svc._lock_repo.list_active_locks()),
    })
